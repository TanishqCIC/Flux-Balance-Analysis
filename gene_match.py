import xlrd
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

file_location_lethal = "/media/tanishqcic/New Volume/class_sem_4/ma_mam/Flux_BA/lethal_genes_0.001.xlsx"
file_location_distinct = "/media/tanishqcic/New Volume/class_sem_4/ma_mam/Flux_BA/pathway_distinct_genes_1.1.xlsx"
file_location_all = "/media/tanishqcic/New Volume/class_sem_4/ma_mam/Flux_BA/trim_1.txt"

workbook = xlrd.open_workbook(file_location_lethal)
sheet = workbook.sheet_by_index(0)

"""Saving lethal genes in list 'lethal_genes' """

lethal_genes = []

for index in range(0, 188):

	gene = sheet.cell_value(index,0)
	lethal_genes.append(gene)
"""
workbook = xlrd.open_workbook(file_location_distinct)
sheet = workbook.sheet_by_index(0)

hike=4

#distinct gene match 

for index in range(0, 122):

	pathway = sheet.cell_value(index, 0)
	length = sheet.cell_value(index, 1)
	length = int(length)
	essential_count = 0

	for alpha_index in range(0, length):

		gene = sheet.cell_value(index, alpha_index+hike)
		print(gene)

		for beta_index in range(0, len(lethal_genes)):

			if gene == lethal_genes[beta_index]:
				essential_count = essential_count+1
				print('essential_gene', gene)
			else:
				essential_count = essential_count
	print(essential_count)
	ws.cell(row= index+1, column=4).value = essential_count
	ws.cell(row=index+1, column=1).value = pathway
	ws.cell(row=index+1, column=6).value = length
	ws.cell(row= index+1, column=3).value = str('found')
	ws.cell(row=index+1, column=5).value = str('lethal out of')
	wb.save("distinct_gene_to_lethal_1.xlsx")
"""
""" multiple gene match """

workbook = xlrd.open_workbook(file_location_all)
sheet = workbook.sheet_by_index(0)

hike=3

for index in range(0, 122):

	print('for pathway', index)

	pathway = sheet.cell_value(index, 0)
	length = sheet.cell_value(index, 1)
	length = int(length)
	essential_count = 0
	print(length)
	print('last value should be', sheet.cell_value(index, length-3))

	for alpha_index in range(0, length):

		print('alpha_index is ', alpha_index)

		gene = sheet.cell_value(index, alpha_index+hike)
		print(gene)

		for beta_index in range(0, len(lethal_genes)):

			if gene == lethal_genes[beta_index]:
				essential_count = essential_count+1
				print('essential_gene', gene)
			else:
				essential_count = essential_count
	print(essential_count)
	ws.cell(row= index+1, column=4).value = essential_count
	ws.cell(row=index+1, column=1).value = pathway
	ws.cell(row=index+1, column=6).value = length
	ws.cell(row= index+1, column=3).value = str('found')
	ws.cell(row=index+1, column=5).value = str('lethal out of')
	wb.save("all_gene_to_lethal.xlsx")
