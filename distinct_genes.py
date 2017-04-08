import xlrd
from openpyxl import Workbook
import json
from pprint import pprint


wb = Workbook()
ws = wb.active

print('Build started')

file_location = "/media/tanishqcic/New Volume/class_sem_4/ma_mam/Flux_BA/trim_1.txt"
#workbook = xlrd.open_workbook(file_location)
#sheet = workbook.sheet_by_index(0)

index=0
data = []


with open('trim_1.txt') as data_file:    
    data_1 = json.load(data_file)

    data.append(data_1)
    pprint(data)
pathway_change=[]
pathway_change.append('1')
pathway_name_old = data_1[1]['pathway name']
for delta_index in range(1,6265):
	print('gene',data_1[delta_index]['gene name'], data_1[delta_index]['pathway name'])
	pathway_name=data_1[delta_index]['pathway name']
	if(pathway_name_old==pathway_name):
		pathway_name_old = pathway_name
	else:
		pathway_change.append(delta_index)
		pathway_name_old = pathway_name
pathway_change.append(6265)
print(pathway_change)
distinct = []
flag=0
for i in range(0, len(pathway_change)):

	pc = int(pathway_change[i])
	if(i+1<len(pathway_change)):
		pc_next = int(pathway_change[i+1])
		print('pcn', pc_next)
	print('pc', pc)
	if pc_next>pc:
		for alpha_index in range(pc, pc_next):
			print(alpha_index)
			cell_gene = data_1[alpha_index]['gene name']
			cell_gene_detail = {'gene name' : data_1[alpha_index]['gene name'], 'pathway name' : data_1[alpha_index]['pathway name']}
			if(alpha_index == pc):
				distinct.append(cell_gene_detail)
			else:
				for beta_index in range(pc, alpha_index):
					#print('chk loop')
					cell_gene_beta = data_1[beta_index]['gene name'] 
					#print(cell_gene_gamma)
					if cell_gene_beta == cell_gene:
						flag=1				
					else:
						#print('value_not_same')
						pass
					
				if flag!=1:
					distinct.append(cell_gene_detail)
				
				#print('col', column, 'row', row)
				
				#print('value saved')
				
				else:
					flag=0
					pass

with open('distinct_1.txt', 'w') as outfile:
		json.dump(distinct, outfile)
"""
for i in range(0,122) :

	distinct = 0

	pathway = sheet.cell_value(i,0)
	print(pathway)
	ws.cell(row=i+1, column=1).value = pathway
	length = sheet.cell_value(i,1)
	#print(length)
	alpha_index = (int(length))
	#print(alpha_index)
	j=4
	flag=0
	for beta_index in range(3, alpha_index-1):
		#print(beta_index)
		cell_gene = sheet.cell_value(i, beta_index)
		#print(cell_gene)
		if beta_index == 3:
			distinct = distinct+1
			ws.cell(row=i+1, column=j+1).value = cell_gene
			#print(cell_gene)
			j=j+1
			wb.save("pathway_distinct_genes_1.1.xlsx")
			#print('value saved')
