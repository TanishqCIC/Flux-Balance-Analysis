import xml.etree.ElementTree
import xlrd
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

index = 1

def parse(fileName, index):

	"""Writing name of .xml file into excel file cell 'A0'"""
	
	ws.cell(row=index, column=1).value = str(fileName[:-4])

	j = 4
	path=''.join(['pathways\\', fileName])
	e = xml.etree.ElementTree.parse(path).getroot()
	
	i=0

	
	for atype in e.findall('entry'):
		leaf = (atype.get('type'))
		
		if(leaf =='gene'):
			gene_name = (atype.get('name'))
			i=i+1
			gene_name = trim(gene_name)
			(len(gene_name))
			(gene_name)
			for beta_index in range(0, len(gene_name)):
				ws.cell(row=index, column=j).value = str(gene_name[beta_index])
				j+=1
			wb.save("pathway_parse_trim_2.1.xlsx")
	print('No. of genes found are ', i)

def trim(gene_name):

	gene = []

	for alpha_index in range(0, len(gene_name)-5):
		if(gene_name[alpha_index]=="m" and gene_name[alpha_index+3]==":"):
			if(len(gene_name)<11):
				gene.append(''.join([gene_name[alpha_index+4], gene_name[alpha_index+5], gene_name[alpha_index+6], gene_name[alpha_index+7], gene_name[alpha_index+8], gene_name[alpha_index+9]]))
			else:
				if(len(gene_name)>(alpha_index+10)):
					
					if(gene_name[alpha_index+10]=="c") :
						gene.append(''.join([gene_name[alpha_index+4], gene_name[alpha_index+5], gene_name[alpha_index+6], gene_name[alpha_index+7], gene_name[alpha_index+8], gene_name[alpha_index+9], gene_name[alpha_index+10]]))
					else:
						gene.append(''.join([gene_name[alpha_index+4], gene_name[alpha_index+5], gene_name[alpha_index+6], gene_name[alpha_index+7], gene_name[alpha_index+8], gene_name[alpha_index+9]]))
				else:
					gene.append(''.join([gene_name[alpha_index+4], gene_name[alpha_index+5], gene_name[alpha_index+6], gene_name[alpha_index+7], gene_name[alpha_index+8], gene_name[alpha_index+9]]))				
	return gene				
